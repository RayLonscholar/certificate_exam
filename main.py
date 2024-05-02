from kivymd.app import MDApp
from kivymd.uix.card import MDCard
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.snackbar import Snackbar, BaseSnackbar
from kivymd.uix.label import MDLabel
from kivymd.uix.button import MDFlatButton
from kivymd.uix.dialog import MDDialog
from kivymd.uix.relativelayout import MDRelativeLayout
from kivy.uix.gridlayout import GridLayout
from kivy.uix.image import Image
from kivy.uix.scrollview import ScrollView
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.button import Button
from kivy.lang import Builder
from kivy.clock import Clock
from kivy.core.window import Window
from kivy.properties import StringProperty, NumericProperty
from kivy.utils import platform
from kivy.metrics import dp
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import webbrowser
import random
import os

d = "SQL.xlsx" # excel路徑
wb = openpyxl.load_workbook(d) # open excel
s1 = wb[f"{wb.sheetnames[0]}"] # 工作區 (default first)
# s1 = wb["Question1"]
req = None # 單字卡選擇項
test_op = 0 # 測驗選項
start = None # 開始題號
end = None # 結束題號
mq = [] # miss questions
sc = "" # scores

# 抓取excel值
def get_values(sheet):
    arr = []                      # 第一層串列
    for row in sheet:
        arr2 = []                 # 第二層串列
        for column in row:
            arr2.append(column.value)  # 寫入內容
        arr.append(arr2)
    return arr
# print(get_values(s1)) # test

# 主選單
class MDLayout(Screen):
    def open_season(self, instance):
        # 開啟工作表選單
        global s1, wb
        self.menu = None
        menu_items = [
            {
                "viewclass" : "OneLineListItem",
                "text" : f"[font=./font/msjh.ttc]{i}[/font]",
                "height" : dp(56),
                "on_release": lambda x=f"{i}": self.open_s1(x),
            } for i in wb.sheetnames
        ]
        if not self.menu:
            self.menu = MDDropdownMenu(
                items=menu_items,
                width_mult=4,
            )
        print(self)
        self.menu.caller = instance
        self.menu.open()
    def open_s1(self, instance):
        global s1, wb
        s1 = wb[instance]
        self.ids.T.title = f"[font=./font/msjh.ttc]Q&A      [{instance}][/font]"
        self.menu.dismiss()
    def open_word(self, instance):
        self.manager.current = "word"
        self.manager.transition.direction="left"
    def open_test(self,instance):
        self.manager.current = "test_op"
        self.manager.transition.direction="left"
    def open_edit(self, instance):
        self.manager.current = "add"
        self.manager.transition.direction="left"
    def open_file(self, instance):
        self.manager.current = "opedit"
        self.manager.transition.direction="left"

class MD3Card(MDCard):
    # 卡片class
    text = StringProperty()

# 背單字介面
class WORDLayout(Screen):
    def create_card(self,args, i):
        index = i # index
        super().on_enter(args)
        self.words = get_values(s1)
        for word in self.words[i:i+50]:
            # display 50 Questions
            self.ids.grid1.add_widget(
                MD3Card(
                    id = f"{index}",
                    line_color = (0.2, 0.2, 0.2, 0.8),
                    style = "outlined",
                    text = f"{word[0]}",
                    padding = "5dp",
                    size_hint = (1, .1),
                    md_bg_color = "#f4dedc",
                    shadow_softness = 2,
                    shadow_offset = (0, 1),
                    on_press = self.annotate
                )
            )
            index += 1

    def on_enter(self, *args):
        self.items = 1
        self.w = [] # All Questions
        # Create Question cards
        super().on_enter(*args)
        self.words = get_values(s1) # All Questions load
        for word in self.words[1:]:
            self.w.append(word)
        self.create_card(self, *args, 1)
        # i = 0 # index
        # for word in words[1:]:
        #     self.w.append(word)
        #     self.ids.grid1.add_widget(
        #         MD3Card(
        #             id = f"{i}",
        #             line_color = (0.2, 0.2, 0.2, 0.8),
        #             style = "outlined",
        #             text = f"{word[0]}",
        #             padding = "5dp",
        #             size_hint = (1, .1),
        #             md_bg_color = "#f4dedc",
        #             shadow_softness = 2,
        #             shadow_offset = (0, 1),
        #             on_press = self.annotate
        #         )
        #     )
        #     i += 1
    def backTop(self,dt):
        self.ids.grid1.height = self.ids.grid1.minimum_height
    
    def page_d(self, instance):
        # print(len(self.w)) # test
        # print(self.items)
        if self.items <= len(self.w)-50:
            # 判斷是否為最後一頁
            self.ids.grid1.clear_widgets()
            self.items += 50
            self.create_card(self, self.items)
            self.ids.grid1.height = 1
            Clock.schedule_once(self.backTop, 1.5) # 返回頂部 1.8s
        else: Question_AlboxApp().underInfo("已經是最後一頁搂")

    def page_u(self, instance):
        if self.items >= 50:
            # 判斷是否為第一頁
            self.ids.grid1.clear_widgets()
            self.items -= 50
            self.create_card(self, self.items)
            self.ids.grid1.height = 1
            Clock.schedule_once(self.backTop, 1.5) # 返回頂部 1.8s
        else: Question_AlboxApp().underInfo("已經是第一頁搂")

    def Info(self, T1, T2, T3, T4, T5, T6, T7, id):
        global req
        # 按鈕會回傳三個值，T1和T2是我們輸出的字
        self.popup = None
        self.id = id
        req = self.id
        self.now_word = f"{T1}"
        box = BoxLayout(orientation = 'vertical', size_hint = (1, 1))
        if not self.popup:
            self.popup = Popup(
                title = f"{T1}",
                title_font = "./font/msjh.ttc",
                title_size= (36),
                title_align = 'center',
                content = box,
                size_hint=(1, 1),
                auto_dismiss=False,
            )
        # print(box.width)
        # print(box.size)
        # print(self.popup.width)
        Img = Image(size_hint = (0, 0), size = (0, 0))
        if T2 != None and T2 != "":
            Img = Image(source = f"{T2}", size_hint = (1, .55))
            print(T2)
        box.add_widget(Img)
        Scroll = ScrollView(size_hint=(1, 0.8), do_scroll_x = False, do_scroll_y = True)
        t = Label(font_name = "./font/msjh.ttc", font_size = 30, text_size = (self.width / 2, None), text = f"題目：{T1}\n\n正確答案：\n    {T3}\n\nA：\n    {T4}\n\nB：\n    {T5}\n\nC：\n    {T6}\n\nD：\n    {T7}")
        t._label.refresh()
        h = t._label.texture.size[1]
        label = Label(size_hint_y = None, height = h, font_name = "./font/msjh.ttc", font_size = 30, text_size = (self.width / 2, None), text = f"題目：{T1}\n\n正確答案：\n    {T3}\n\nA：\n    {T4}\n\nB：\n    {T5}\n\nC：\n    {T6}\n\nD：\n    {T7}")
        Scroll.add_widget(label)
        print(label.width)
        print(label.height)
        box.add_widget(Scroll)
        box2 = BoxLayout(orientation = 'horizontal', padding = (2), size_hint = (1,0.1))
        box2.add_widget(Button(text = "DIS", font_size = 35, on_press=self.popup.dismiss))
        box2.add_widget(Button(text = "Edit", font_size = 35, on_press=self.edit))
        box.add_widget(box2)
        Question_AlboxApp().theme_cls.theme_style = "Dark" # window reset color
        self.popup.open()

    def edit(self, instance):
        # delete Question
        # print(self.id)
        # s1.delete_rows(self.id) # 刪除row
        # wb.save("English.xlsx")
        self.popup.dismiss()
        self.ids.grid1.clear_widgets()
        self.manager.current = "edit"
        self.manager.transition.direction="left"
        # 刷新頁面
        # self.ids.grid1.clear_widgets()
        # self.create_card(self, self.items)

    def annotate(self, instance):
        # Question information
        # print(int(instance.id)) # test
        print(self.w[int(instance.id)-1][2])
        self.Info(instance.text, self.w[int(instance.id)-1][1], self.w[int(instance.id)-1][2], self.w[int(instance.id)-1][3], self.w[int(instance.id)-1][4], self.w[int(instance.id)-1][5], self.w[int(instance.id)-1][6], int(instance.id)+1)
        
    def back(self, instance):
        self.ids.grid1.clear_widgets()
        self.manager.current = "OP"
        self.manager.transition.direction="right"

class TEST_OPLayout(Screen):
    # 選擇測驗
    def ck_items(self, start, end):
        global test_op, s1
        t = []
        t_items = get_values(s1) # All Questions load
        for word in t_items[1:]:
            t.append(word)
        if 0 < start <= end and 0 < end <= len(t) :
            print("pass")
            if test_op == 1:
                self.manager.current = "mc"
                self.manager.transition.direction="left"
            if test_op == 2:
                self.manager.current = "cloze"
                self.manager.transition.direction="left"
        else: Question_AlboxApp().HintInfo("Error", "輸入題號有錯誤(請確認開始和結束題號)!")

    def q_all(self, instance):
        global s1
        t = []
        t_items = get_values(s1) # All Questions load
        for word in t_items[1:]:
            t.append(word)
        self.ids.start.text = "1"
        self.ids.end.text = f"{len(t)}"

    def ch_test(self, instance):
        # 選擇題
        global test_op, start, end
        try:
            test_op = 1
            start = int(self.ids.start.text)
            end = int(self.ids.end.text)
            self.ck_items(start, end)
        except ValueError: Question_AlboxApp().HintInfo("Error", "輸入題號有錯誤(請確認開始和結束題號)!")

    def cloze(self, instance):
        # 填充題
        global test_op, start, end
        try:
            test_op = 2
            start = int(self.ids.start.text)
            end = int(self.ids.end.text)
            self.ck_items(start, end)
        except ValueError: Question_AlboxApp().HintInfo("Error", "輸入題號有錯誤(請確認開始和結束題號)!")
    
    def back(self, instance):
        self.ids.start.text = ""
        self.ids.end.text = ""
        self.manager.current = "OP"
        self.manager.transition.direction="right"

class Multiple_choiceLayout(Screen):
    # 選擇題
    def test_word(self, word):
        global start, end, mq, sc
        try:
            self.qst = random.sample(word, 1)[0] # Once Questions load
            self.w.remove(self.qst) # 刪除選過的題目
            dic = {"A":f"{self.qst[3]}","B":f"{self.qst[4]}","C":f"{self.qst[5]}","D":f"{self.qst[6]}"}
            if self.qst[1] != None:
                self.ids.q_image.source = f"{self.qst[1]}"
                self.ids.q_image.size_hint = (1, .65)
            else:
                self.ids.q_image.source = ""
                self.ids.q_image.size = 0, 0
                self.ids.q_image.size_hint = (0, 0)
            self.n = self.qst[0] # question name
            self.ans = dic[self.qst[2]] # answer
            self.ran_w = random.sample(self.qst[3:], 4) # random 4個選項
            t = Label(font_name = "./font/msjh.ttc", font_size = 30, text_size = (self.width / 2, None), text = f"題目：{self.n}\n\nA：\n    {self.ran_w[0]}\n\nB：\n    {self.ran_w[1]}\n\nC：\n    {self.ran_w[2]}\n\nD：\n    {self.ran_w[3]}")
            t._label.refresh()
            self.ids.test_word.height = t._label.texture.size[1]
            # print(word)
            # print(self.qst)
            print(self.ans)
            # print(self.ran_w)
            self.ids.test_word.text = f"題目：{self.n}\n\nA：\n    {self.ran_w[0]}\n\nB：\n    {self.ran_w[1]}\n\nC：\n    {self.ran_w[2]}\n\nD：\n    {self.ran_w[3]}"
            # self.ids.one.text = f"{self.ran_w[0]}"
            # self.ids.two.text = f"{self.ran_w[1]}"
            # self.ids.three.text = f"{self.ran_w[2]}"
            # self.ids.four.text = f"{self.ran_w[3]}"
        except ValueError:
            Question_AlboxApp().HintInfo("測驗完畢", "")
            sc = f"{end-len(mq)} / {end-start+1}"
            self.manager.current = "check"
            self.manager.transition.direction="left"
        except KeyError:
            Question_AlboxApp().HintInfo("Error", "請確認題目的正確答案是否為ABCD")
            self.manager.current = "test_op"
            self.manager.transition.direction="right"
    
    def save_miss(self, quest):
        global mq
        if quest not in mq:
            mq.append(quest)
        # print(mq)

    def on_enter(self, *args):
        # Reset question
        global test_op, start, end, mq, sc
        self.a = 0
        self.i = 1
        self.w = []
        sc = "" # scores
        mq = [] # miss questions
        self.words = get_values(s1) # All Questions load
        for word in self.words[int(start):int(end)+1]:
            self.w.append(word)
        print(self.w)
        self.test_word(self.w)

    def test(self, instance):
        # 判斷答案是否正確
        # print(instance.text)
        dic = {"A":f"{self.ran_w[0]}","B":f"{self.ran_w[1]}","C":f"{self.ran_w[2]}","D":f"{self.ran_w[3]}"}
        if dic[instance.text] == self.ans:
            self.test_word(self.w)
        else: 
            self.save_miss(self.qst)
            Question_AlboxApp().underInfo("答錯了，在試一次")

    def back(self, instance):
        self.manager.current = "test_op"
        self.manager.transition.direction="right"

class Check_miss(Screen):
    # 查看分數結果
    global mq, sc
    def show(self, quest):
        if quest[1] != None:
                self.ids.q_image.source = f"{quest[1]}"
                self.ids.q_image.size_hint = (1, .65)
        else: 
            self.ids.q_image.source = ""
            self.ids.q_image.size = 0, 0
            self.ids.q_image.size_hint = (0, 0)
        t = Label(font_name = "./font/msjh.ttc", font_size = 30, text_size = (self.width / 2, None), text = f"題目：{quest[0]}\n\n正確答案：\n    {quest[2]}\n\nA：\n    {quest[3]}\n\nB：\n    {quest[4]}\n\nC：\n    {quest[5]}\n\nD：\n    {quest[6]}")
        t._label.refresh()
        self.ids.miss_q.height = t._label.texture.size[1]
        self.ids.miss_q.text = f"題目：{quest[0]}\n\n正確答案：\n    {quest[2]}\n\nA：\n    {quest[3]}\n\nB：\n    {quest[4]}\n\nC：\n    {quest[5]}\n\nD：\n    {quest[6]}"

    def on_enter(self, *args):
        # print(mq)
        self.i = len(mq)
        self.now = 0
        self.ids.score.text = f"分數：{sc}分"
        self.ids.b.disabled = True
        self.ids.n.disabled = True
        if mq != []:
            self.ids.b.disabled = False
            self.ids.n.disabled = False
            self.show(mq[0])
            Question_AlboxApp().underInfo("請再複習您答錯的題目")
        else: Question_AlboxApp().underInfo("恭喜您，全部答對")

    def to_down(self, instance):
        if self.now != self.i - 1:
            self.now = self.now + 1
            # print(self.now)
            self.show(mq[self.now])
        else: Question_AlboxApp().underInfo("已經是最後一頁搂")

    def to_up(self, instance):
        if self.now != 0:
            self.now = self.now - 1
            # print(self.now)
            self.show(mq[self.now])
        else: Question_AlboxApp().underInfo("已經是第一頁搂")

    def back(self, instance):
        self.ids.miss_q.text = ""
        self.manager.current = "test_op"
        self.manager.transition.direction="right"

class ClozeLayout(Screen):
    # 填充題
    def test_word(self, word):
        global start, end, mq, sc
        try:
            self.qst = random.sample(word, 1)[0] # Once Questions load
            self.w.remove(self.qst) # 刪除選過的題目
            dic = {"A":f"{self.qst[3]}","B":f"{self.qst[4]}","C":f"{self.qst[5]}","D":f"{self.qst[6]}"}
            if self.qst[1] != None:
                self.ids.q_image.source = f"{self.qst[1]}"
                self.ids.q_image.size_hint = (1, .65)
            else:
                self.ids.q_image.source = ""
                self.ids.q_image.size = 0, 0
                self.ids.q_image.size_hint = (0, 0)
            self.n = self.qst[0] # question name
            self.ans = dic[self.qst[2]].strip() # answer
            self.ran_w = random.sample(self.qst[3:], 4) # random 4個選項
            t = Label(font_name = "./font/msjh.ttc", font_size = 30, text_size = (self.width / 2, None), text = f"題目：{self.n}\n\nA：\n    {self.ran_w[0]}\n\nB：\n    {self.ran_w[1]}\n\nC：\n    {self.ran_w[2]}\n\nD：\n    {self.ran_w[3]}")
            t._label.refresh()
            self.ids.test_word.height = t._label.texture.size[1]
            # print(word)
            # print(self.qst)
            # print(self.ans)
            # print(self.ran_w)
            self.ids.test_word.text = f"題目：{self.n}\n\nA：\n    {self.ran_w[0]}\n\nB：\n    {self.ran_w[1]}\n\nC：\n    {self.ran_w[2]}\n\nD：\n    {self.ran_w[3]}"
            # self.ids.one.text = f"{self.ran_w[0]}"
            # self.ids.two.text = f"{self.ran_w[1]}"
            # self.ids.three.text = f"{self.ran_w[2]}"
            # self.ids.four.text = f"{self.ran_w[3]}"
        except ValueError:
            Question_AlboxApp().HintInfo("測驗完畢", "")
            sc = f"{end-len(mq)} / {end-start+1}"
            self.manager.current = "check"
            self.manager.transition.direction="left"
        except KeyError:
            Question_AlboxApp().HintInfo("Error", "請確認題目的正確答案是否為ABCD")
            self.manager.current = "test_op"
            self.manager.transition.direction="right"
    
    def save_miss(self, quest):
        global mq
        if quest not in mq:
            mq.append(quest)
        # print(mq)

    def on_enter(self, *args):
        # Reset question
        global test_op, start, end, mq, sc
        self.w = []
        sc = "" # scores
        mq = [] # miss questions
        self.words = get_values(s1) # All Questions load
        for word in self.words[int(start):int(end)+1]:
            self.w.append(word)
        print(self.w)
        self.test_word(self.w)

    def test(self, instance):
        # 判斷答案是否正確
        if self.ids.cloze_input.text.strip() == self.ans:
            self.test_word(self.w)
        else: 
            self.save_miss(self.qst)
            Question_AlboxApp().underInfo("答錯了，在試一次")

    def back(self, instance):
        self.ids.cloze_input.text = ""
        self.manager.current = "test_op"
        self.manager.transition.direction="right"

class OPEDITLayout(Screen):
    # 選擇修改工作表
    def open_work(self, instance):
        self.manager.current = "edwork"
        self.manager.transition.direction="left"

    def open_copy(self, instance):
        self.manager.current = "openfile"
        self.manager.transition.direction="left"

    def back(self, instance):
        self.manager.current = "OP"
        self.manager.transition.direction="right"

class EDITLayout(Screen):
    # 修改內容
    def on_enter(self, *args):
        global req
        # print(req)
        self.text = "" # 指令內容
        i = 0 # 指令起始位置
        self.w = [] # all Questions
        self.words = get_values(s1) # All Questions load
        for word in self.words[1:]:
            self.w.append(word)
        # print(self.w) # test
        edit = self.w[req-2]
        for info in edit:
            if info != None and i != 0:
                # print(info) # test
                self.text += f"!n{info}"
            if i == 0:
                i = 1
                self.text += f"{info}"
        print(self.text)
        self.ids.edit_label.text = f"{self.w[req-2][0]}"
        self.ids.edit_input.text = f"{self.text}"

    def back(self, instance):
        self.ids.edit_label.text = ""
        self.ids.edit_input.text = ""
        self.manager.current = "word"
        self.manager.transition.direction="right"

    def delete(self, instance):
        global d
        # 刪除單字
        # print(req)
        s1.delete_rows(req) # 刪除row
        wb.save(d)
        self.ids.edit_label.text = ""
        self.ids.edit_input.text = ""
        self.manager.current = "word"
        self.manager.transition.direction="right"

    def submit(self, instance):
        # 提交修改
        global req, d
        # print(req) # test
        result = self.ids.edit_input.text.split("!n")
        # print(result) # test
        col = 1
        self.ids.edit_label.text = f"{result[0]}"
        s1.delete_rows(req) # 刪除
        s1.insert_rows(req) # 創建空列
        for r in result:
            print(r)
            s1.cell(req, col).value = r # 輸入值
            col += 1
        wb.save(d)

class ADDLayout(Screen):
    # 新增Excel
    def Add(self, instance):
        global d
        w = [] # all words
        p = self.ids.Input.text
        if p != "":
            # 判斷是否為空值
            result = p.split("!n")
            print(result)
            words = get_values(s1) # All Questions load
            try:
                for word in words[1:]:
                    w.append(word[0].strip().lower())
            except AttributeError:
                Question_AlboxApp().HintInfo("Error", "表格有空值!")
            # print(w)
            print("新增成功")
            row = s1.max_row + 1
            col = 1
            for r in result:
                s1.cell(row, col).value = r
                s1.cell(row, col).font = Font(name = "Microsoft JhengHei", size = 8) # 字體
                s1.cell(row, col).alignment = Alignment(vertical = "center") # 對齊
                col += 1
            wb.save(d)
            Question_AlboxApp().HintInfo(f"{result[0]}", "新增成功")
            self.ids.Input.text = ""
        else: Question_AlboxApp().HintInfo("Error", "請輸入指令!")

    def back(self, instance):
        self.ids.Input.text = ""
        self.manager.current = "OP"
        self.manager.transition.direction="right"

class EDworkLayout(Screen):
    # 新增刪除工作表
    def add_w(self, instance):
        global wb, d, s1
        text = self.ids.work1.text # 讀取textinput
        if text != "":
            if text not in wb.sheetnames:
                c = wb["列表格式"]
                copy = wb.copy_worksheet(c) # copy工作表
                copy.title = text
                wb.move_sheet(copy, -1)
                wb.save(d) # save
                s1 = wb[text]
                self.ids.work1.text = ""
                Question_AlboxApp().HintInfo("新增成功", "")
                self.manager.get_screen("OP").ids.T.title = f"[font=./font/msjh.ttc]Q&A      [{text}][/font]"
            else: Question_AlboxApp().HintInfo("Error", "重複的工作表名稱")
        else: Question_AlboxApp().HintInfo("Error", "請輸入要新增工作表名稱")

    def dlt_w(self, instance):
        global wb, d, s1
        text = self.ids.work2.text # 讀取textinput
        if text != "":
            if text in wb.sheetnames:
                dlt = wb[text]
                wb.remove_sheet(dlt) # 刪除工作表
                wb.save(d) # save
                s1 = wb[f"{wb.sheetnames[0]}"]
                self.ids.work2.text = ""
                Question_AlboxApp().HintInfo("刪除成功", "")
                self.manager.get_screen("OP").ids.T.title = f"[font=./font/msjh.ttc]Q&A      [{wb.sheetnames[0]}][/font]"
            else: Question_AlboxApp().HintInfo("Error", "請確認工作表名稱")
        else: Question_AlboxApp().HintInfo("Error", "請輸入要刪除工作表名稱")

    def back(self, instance):
        self.manager.current = "opedit"
        self.manager.transition.direction="right"

class OPENLayout(Screen):
    # 開啟excel舊檔
    def on_enter(self, *args):
        self.input = None

    def selected(self,filename):
        try:
            print(filename[0])
            self.dir = filename[0]
            self.sour = os.path.splitext(self.dir) # 抓取副檔名
            print(self.dir)
            if self.sour[1] == ".xlsx":
                self.input = 1
                print("選取成功")
            else:
                self.input = None
        except:
            pass
    
    def f_open(self, instance):
        try:
            PATH ="."
            if platform == "android":
                from android.permissions import request_permissions, Permission
                request_permissions([Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE])
                app_folder = os.path.dirname(os.path.abspath(__file__))
                PATH = "/storage/emulated/0" #app_folder
            op = openpyxl.load_workbook(f"{self.dir}")
            o1 = op[f"{op.sheetnames[0]}"] # open_excel(第一個)工作表內容
            self.n = [] # new words
            self.word1s = get_values(o1) # 讀取所有單字
            for word1 in self.word1s[1:]:
                self.n.append(word1)
            print(self.n)
            word2s = get_values(s1) # 讀取所有單字
            for word2 in word2s[1:]:
                s1.delete_rows(2) # 刪除舊的單字列
            row = 2
            col = 1
            for r in self.n:
                for c in r:
                    # 創建新列
                    s1.cell(row, col).value = c
                    s1.cell(row, col).font = Font(name = "Microsoft JhengHei", size = 8) # 字體
                    s1.cell(row, col).alignment = Alignment(vertical = "center") # 對齊
                    col += 1
                col = 1
                row += 1
            wb.save(d)
            Question_AlboxApp().HintInfo(f"新增成功", "")
        except: Question_AlboxApp().HintInfo("Error", "請選擇excel檔案")

    def back(self, instance):
        self.manager.current = "opedit"
        self.manager.transition.direction="right"

class WindowManager(ScreenManager):
    pass

class CustomSnackbar(BaseSnackbar):
    text = StringProperty(None) # 文字
    icon = StringProperty(None) # 圖示
    font_size = NumericProperty("30sp")
    duration = 0.5 # 持續時間(seconds)

class Question_AlboxApp(MDApp):
    popup = None
    def build(self):
        self.theme_cls.theme_style = "Dark" # window reset color
        self.icon = "img.ico"
        kv = Builder.load_file("main2.kv") # load kivy file
        return kv

    def HintInfo(self, T1, T2):
        # 按鈕會回傳兩個值，T1和T2是我們輸出的字
        self.now_word = f"{T1}"
        box = BoxLayout(orientation = 'vertical')
        if not self.popup:
            self.popup = Popup(
                title = f"{T1}",
                title_font = "./font/msjh.ttc",
                title_size= (60),
                title_align = 'center',
                content = box,
                size_hint=(0.5, 0.5),
                auto_dismiss=False,
            )
        label = Label(size_hint = (1,1), text = f"{T2}", text_size = (Window.width / 2.5,None), font_name = "./font/msjh.ttc", font_size = 40)
        box.add_widget(label)
        box2 = BoxLayout(orientation = 'horizontal', padding = (2), size_hint = (1,0.4))
        box2.add_widget(Button(text = "DIS", font_size = 40, on_press=self.popup.dismiss))
        box.add_widget(box2)
        self.theme_cls.theme_style = "Dark" # window reset color
        self.popup.open()

    def underInfo(self, T1):
        # 按鈕會回傳兩個值，T1和T2是我們輸出的字
        snackbar = CustomSnackbar(
            text=f"{T1}",
            icon="information",
            snackbar_x = "10dp",
            snackbar_y = f"{Window.height / 10}dp",
        )
        snackbar.size_hint_x = (
            Window.width - (snackbar.snackbar_x * 2)
        ) / Window.width
        self.theme_cls.theme_style = "Dark" # window reset color
        snackbar.open()

if __name__ == "__main__":
    Question_AlboxApp().run()

# 新增tag單字