import docx
import openpyxl
import os
import re

def process_word_file(selected_file):
    filename = f"{os.path.splitext(selected_file)[0].split('/')[-1]}" # 題庫word名稱
    doc = docx.Document(selected_file) # 開啟word
    para = doc.paragraphs # 整份文件內容
    # print('段落數量： ', len(para),'\n')
    topic_picture_location = [] # 題目圖片位置，["", "<img1>", "", "<img2>"]
    topic_picture_index = 0 # 題目圖片數量
    index = 0 # 題號
    is_first_topic = False # 是否為題目開始的段落
    is_topic = False # 是否為題目
    answer = "" # 答案
    save_path = f"./assets/imgs/{filename}" # 圖片儲存位址
    memory = []

    # 匯出word裡的圖片
    def w_img(blob_data, save_path):
        with open(save_path, "wb") as f:
            f.write(blob_data)

    # 資料傳入excel
    def to_excel(workbook = filename):
        excel = openpyxl.load_workbook("exam_data.xlsx") # 讀取excel檔
        all_sheetnames = excel.sheetnames
        if workbook in all_sheetnames: # 判斷輸入的工作表是否存在
            wb = excel[workbook] # 開啟工作表
        else:
            # 創建工作表並開啟
            excel.create_sheet("{}".format(workbook))
            wb = excel[workbook]
        for column_index, column_content in enumerate(memory):
            for row_index, row_content in enumerate(column_content):
                wb.cell(column_index+1, row_index+1).value = row_content
        excel.save("exam_data.xlsx")

    # 取得word裡的題目(含選項)與圖片
    for _ in range(0, len(para)):
        try:
            content = para[_] # 每個段落內容
            no_ans_topic = "" # 去掉答案的題目
            img_info = [""] # 圖片資訊
            option = [] # 選項
            if content.text != "": # 判斷文本是否為空值
                # print(f"原：{content.text}")
                no_num_topic = content.text[re.search(r"\D", content.text).start():].strip() # 去掉題目的題號
                # print(f"修改後：{no_num_topic}")
                no_num_topic = re.sub(r'\(', '（', no_num_topic) # 修正括弧為全形
                no_num_topic = re.sub(r'\)', '）', no_num_topic) # 修正括弧為全形
                compare = re.search(r'（(.*?)）：|\((.*?)\)：', no_num_topic) # 尋找指定規則
                if compare: # 取得題目的答案並去除
                    # print(f"compare：{compare}")
                    answer = compare.groups() # 取得答案
                    no_ans_topic = re.sub(r'（(.*?)）：|\((.*?)\)：', '( )：', no_num_topic) # 去掉答案的題目
                if re.match(r"^[A-D]:|^[A-D]：", content.text):
                    option = [content.text[0], content.text[1]] # 選項
                # 取得題目和答案
                if len(no_ans_topic) > 4:
                    if [no_ans_topic[0], no_ans_topic[2], no_ans_topic[3]] == ['(', ')', '：']:
                        topic_picture_location = [] # 重置題目圖片位置
                        topic_picture_index = 0 # 重置題目圖片數量
                        memory.append([])
                        index += 1 # 題號
                        is_first_topic = True # 是否為題目的第一個段落
                        is_topic = True # 是否為題目
                        print(index, no_ans_topic)
                        memory[index-1].insert(0, no_ans_topic)
                        memory[index-1].insert(1, f"{topic_picture_location}")
                        for _ in answer:
                            if _:
                                memory[index-1].insert(2, _.strip())

                # match ... case ... 只支援python 3.10以上
                # 為了相容性則不選擇使用      
                if option == ['A', '：']: # 選項A
                    is_topic = False
                    img_info = ["A", 3]
                    print(content.text)
                    memory[index-1].insert(img_info[1], content.text)
                elif option == ['B', '：']: # 選項B
                    is_topic = False
                    img_info = ["B", 4]
                    print(content.text)
                    memory[index-1].insert(img_info[1], content.text)
                elif option == ['C', '：']: # 選項C
                    is_topic = False
                    img_info = ["C", 5]
                    print(content.text)
                    memory[index-1].insert(img_info[1], content.text)
                elif option == ['D', '：']: # 選項D
                    is_topic = False
                    img_info = ["D", 6]
                    print(content.text)
                    memory[index-1].insert(img_info[1], content.text)
                if is_topic and is_first_topic == False: # 判斷是否還有題目
                    print(content.text)
                    topic_picture_location.append("") # 題目圖片位置
                    memory[index-1][0] = "{}{}".format(memory[index-1][0], "!n{}".format(content.text))
                is_first_topic = False
            imgs = content._element.xpath(".//pic:pic")
            if imgs:
                for img in imgs:
                    print(img)
                    rel = img.xpath('.//a:blip/@r:embed')[0] # 取得picture's path
                    # print(rel)
                    image_part = doc.part.related_parts[rel]
                    img_blob = image_part.image.blob # 轉為二進制
                    if img_info != [""]: # 判斷是否為選項圖片
                        path = "{}img{}{}.png".format(save_path, index, img_info[0])
                        memory[index-1][img_info[1]] = "{}".format(f"{path}")
                        w_img(img_blob, path) # 下載image
                    else:
                        topic_picture_index+= 1
                        path = "{}img{}_{}.png".format(save_path, index, topic_picture_index)
                        topic_picture_location.append(path) # 題目圖片位置
                        memory[index-1][0] = "{}{}".format(memory[index-1][0], "!n<img>") # 題目圖片位置
                        w_img(img_blob, path) # 下載image
                        # 新增選項圖片(同理)
                        # print(f"{topic_picture_location}")
            memory[index-1][1] = f"{topic_picture_location}"

        except IndexError:
            print("IndexERROR")
    # print(memory)
    to_excel()