import tkinter as tk
from tkinter import filedialog, messagebox
import docx
import openpyxl
import os
import re

def process_word_file(selected_file, save_images_path):
    filename = f"{selected_file}" # 題庫word名稱
    doc = docx.Document(f"./questions/{filename}.docx") # 開啟word
    para = doc.paragraphs # 整份文件內容
    # print('段落數量： ', len(para),'\n')
    index = 0 # 題號
    is_first_topic = False # 是否為題目開始的段落
    is_topic = False # 是否為題目
    answer = "" # 答案
    save_path = f"{save_images_path}/"+filename # 圖片儲存位址

    # 匯出word裡的圖片
    def w_img(blob_data, save_path):
        with open(save_path, "wb") as f:
            f.write(blob_data)

    # word資料傳入excel
    def to_excel(workbook, data, x = "1", y = "A", is_topic = None):
        excel = openpyxl.load_workbook("exam_data.xlsx") # 讀取excel檔
        all_sheetnames = excel.sheetnames
        if workbook in all_sheetnames: # 判斷輸入的工作表是否存在
            wb = excel[workbook] # 開啟工作表
        else:
            # 創建工作表並開啟
            excel.create_sheet("{}".format(workbook))
            wb = excel[workbook]
        # 寫入資料至excel
        if is_topic == None:
            wb["{}{}".format(y, x)].value = data
        else:
            wb["{}{}".format(y, x)].value = "{}{}".format(wb["{}{}".format(y, x)].value, data)

        excel.save("exam_data.xlsx")

    # 取得word裡的題目(含選項)與圖片
    for _ in range(0, len(para)):
        try:
            content = para[_] # 每個段落內容
            no_ans_topic = "" # 去掉答案的題目
            img_info = [""] # 圖片資訊
            if content.text != "": # 判斷文本是否為空值
                no_num_topic = re.sub(r'[0-9]', '', content.text) # 去掉題目的題號
                compare = re.search(r'（(.*?)）：|\((.*?)\)：', no_num_topic) # 尋找指定規則
                if compare: # 取得題目的答案並去除
                    answer = compare.groups() # 取得答案
                    no_ans_topic = re.sub(r'（(.*?)）：|\((.*?)\)：', '( )：', no_num_topic) # 去掉答案的題目
                option = [content.text[0], content.text[1]] # 選項
                # 取得題目和答案
                if len(no_ans_topic) > 4:
                    if [no_ans_topic[0], no_ans_topic[2], no_ans_topic[3]] == ['(', ')', '：']:
                        index += 1
                        is_first_topic = True # 是否為題目的第一個段落
                        is_topic = True # 是否為題目
                        print(index, no_ans_topic)
                        to_excel(filename, no_ans_topic, index, "A") # 傳入題目
                        for _ in answer:
                            if _:
                                to_excel(filename, _, index, "C") # 傳入答案

                # match ... case ... 只支援python 3.10以上
                # 為了相容性則不選擇使用      
                if option == ['A', '：']: # 選項A
                    is_topic = False
                    img_info = ["A", "D"]
                    print(content.text)
                    to_excel(filename, content.text, index, "D") # 傳入選項
                elif option == ['B', '：']: # 選項B
                    is_topic = False
                    img_info = ["B", "E"]
                    print(content.text)
                    to_excel(filename, content.text, index, "E") # 傳入選項
                elif option == ['C', '：']: # 選項C
                    is_topic = False
                    img_info = ["C", "F"]
                    print(content.text)
                    to_excel(filename, content.text, index, "F") # 傳入選項
                elif option == ['D', '：']: # 選項D
                    is_topic = False
                    img_info = ["D", "G"]
                    print(content.text)
                    to_excel(filename, content.text, index, "G") # 傳入選項
                if is_topic and is_first_topic == False: # 判斷是否還有題目
                    # 待新增判斷是否還有題目圖片
                    print(content.text)
                    to_excel(filename, "!n{}".format(content.text), index, "A", 1)
                is_first_topic = False
            img = content._element.xpath(".//pic:pic")
            if img:
                print(img)

                img = img[0]
                rel = img.xpath('.//a:blip/@r:embed')[0] # 取得picture's path
                image_part = doc.part.related_parts[rel]
                img_blob = image_part.image.blob # 轉為二進制
                path = "{}img{}{}.png".format(save_path, index, img_info[0])
                w_img(img_blob, path) # 下載image
                if img_info != [""]: # 判斷是否為選項圖片
                    to_excel(filename, path, index, img_info[1]) # 傳入選項圖片的位址
                else:
                    to_excel(filename, path, index, "B") # 傳入題目圖片的位址

        except IndexError:
            print("IndexERROR")

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("Word files", "*.docx")])
    if file_path:
        file_label.config(text=f"Selected file: {file_path}", wraplength = int(app.winfo_width()/1.2))
        app.selected_file = os.path.splitext(file_path)[0].split('/')[-1]

def select_save_path():
    path = filedialog.askdirectory()
    if path:
        path_label.config(text=f"Save images to: {path}", wraplength = int(app.winfo_width()/1.2))
        app.save_images_path = path

def start_processing():
    if hasattr(app, 'selected_file') and hasattr(app, 'save_images_path'):
        process_word_file(app.selected_file, app.save_images_path)
        messagebox.showinfo("Successfully", "Processing completed and data saved to exam_data.xlsx.\n資料儲存到exam_data.xlsx")
    else:
        messagebox.showwarning("Input Required", "Please select both a Word file and a saving path for images.\n請選擇word檔和照片存放位置")

app = tk.Tk()
app.title("Word to Excel User Interface")
window_width = 500
window_height = 300
app.geometry(f'{window_width}x{window_height}')

# 檔案選擇按鈕
file_btn = tk.Button(app, text="請選擇word檔案", font = ('Arial',16), command=select_file)
file_btn.pack(pady=5)
file_label = tk.Label(app, text="請選擇", font = ('Arial',10))
file_label.pack(pady=5)

# 圖片儲存路徑選擇按鈕
path_btn = tk.Button(app, text="選擇照片存放位置", font = ('Arial',16), command=select_save_path)
path_btn.pack(pady=5)
path_label = tk.Label(app, text="請選擇", font = ('Arial',10))
path_label.pack(pady=5)

# 開始處理按鈕
process_btn = tk.Button(app, text="開始轉換", font = ('Arial',16), command=start_processing)
process_btn.pack(pady=20)

app.mainloop()
