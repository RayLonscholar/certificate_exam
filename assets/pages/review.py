from flet import *
from flet_route import Params, Basket
from assets.pages.navigation import rail_item
import openpyxl

class Review():
    def __init__(self, page):
        # super().__init__()

        # 提示對話
        def warning_info(text:str):
            info = AlertDialog(
                title=Text(text)
            )
            page.dialog = info
            info.open = True
            page.update()

        def restart_clicked(e):
            self.choice.visible = True # 顯示第一區塊
            self.topic.visible = False # 隱藏第二區塊
            self.topic_information_main.visible = False # 隱藏第三區塊
            self.edit_content.visible = False # 隱藏修改內容按鍵
            self.topic_edit_main.visible = False # 隱藏第四區塊
            write_list() # 刷新工作表選擇
            page.update()

        def show_edit(e):
            self.topic_information_main.visible = False # 隱藏第三區塊
            self.topic_edit_main.visible = True # 顯示第四區塊
            page.update()

        # 重新選擇題庫
        self.restart = TextButton(
            content = Container(
                content=Column(
                    controls = [
                        Text(value="重新選擇題庫", size=20),
                    ],
                    alignment=MainAxisAlignment.CENTER,
                    spacing=5,
                ),
            ),
            on_click=restart_clicked)
        
        # 修改內容
        self.edit_content = TextButton(
            visible = False,
            content = Container(
                content=Column(
                    controls = [
                        Text(value="修改內容", size=20),
                    ],
                    alignment=MainAxisAlignment.CENTER,
                    spacing=5,
                ),
            ),
            on_click=show_edit
        )

        def review_submit_clicked(e):
            # 讀取excel工作表裡的每一行
            def get_values(sheet):
                arr = []
                # 行
                for row in sheet:
                    arr2 = []
                    # 列
                    for column in row:
                        arr2.append(column.value) # 寫入內容
                    arr.append(arr2)
                return arr
            # 創建題目內容
            def create_topic_information(content):
                # print(content)
                # 清空元件
                self.topic_information.controls = []
                self.topic_up_down_button.controls = []
                self.topic_edit.controls = []
                self.topic_edit_button.controls = []

                # 代修改的值
                self.topic_text_edit = TextField(label="題目", value = "", text_size= 20)
                self.topic_img_edit = TextField(label="題目圖片", value = "", text_size= 20)
                self.topic_answer_edit = TextField(label="答案", value = "", text_size= 20)
                self.topic_option1_edit = TextField(label="選項1", value = "", text_size= 20)
                self.topic_option2_edit = TextField(label="選項2", value = "", text_size= 20)
                self.topic_option3_edit = TextField(label="選項3", value = "", text_size= 20)
                self.topic_option4_edit = TextField(label="選項4", value = "", text_size= 20)

                # 題目
                order = content[0].split("!n")
                self.topic_text_edit.value = f"{content[0]}"
                self.topic_answer_edit.value = f"{content[2]}"
                self.topic_information.controls.append(
                    Text("{}.({})：{}".format(self.index, content[2], order[0][4:]), size = 20, weight = "bold", selectable = True)
                )
                # 題目圖片
                # print(eval(content[1]))
                self.topic_img_edit.value = f"{content[1]}"
                try:
                    for index, _ in enumerate(eval(content[1])):
                        if _ != "": # 有題目圖片
                            self.topic_information.controls.append(
                                Image(src=f"{_}")
                            )
                        else: # 沒有題目圖片
                            self.topic_information.controls.append(
                                Text(f"{order[index+1]}", size = 20, weight = "bold", selectable = True)
                            )
                except TypeError:
                    # print("TypeERROR")
                    warning_info("發生錯誤：請確認題庫資料格式是否正確\n錯誤代碼：0x03")
                    restart_clicked(self) # 重新選擇題庫

                self.topic_information.controls.append(
                    Text("", size = 20, weight = "bold")
                )

                # 選項
                for index, _ in enumerate(content[3:]):
                    option = ['A', 'B', 'C', 'D']
                    if _ != None:
                        if _[0] in option: # 判斷是否有圖片
                            self.topic_information.controls.append(
                                Text(f"{_}", size = 20, weight = "bold")
                            )
                        else:
                            self.topic_information.controls.append(
                                Row(
                                    controls = [
                                        Text(f"{option[index]}：", size = 20, weight = "bold"),
                                        Image(src=f"{_}")
                                    ],
                                )
                            )
                for index, _ in enumerate([self.topic_option1_edit, self.topic_option2_edit, self.topic_option3_edit, self.topic_option4_edit]):
                    try:
                        _.value = content[3+index]
                    except:
                        _.value = ""
                # 上一題、下一題、返回按鈕
                def topic_up_down_button_clicked(e):
                    if e.control.data == "上一題":
                        self.index -= 1
                        content = self.topic_data[self.index-1] # 讀取題目內容
                        create_topic_information(content)
                    if e.control.data == "下一題":
                        self.index += 1
                        content = self.topic_data[self.index-1] # 讀取題目內容
                        create_topic_information(content)
                    if e.control.data == "返回":
                        self.topic.visible = True # 顯示第二區塊
                        self.topic_information_main.visible = False # 隱藏第三區塊
                        self.edit_content.visible = False # 隱藏修改內容選項
                    page.update()

                self.topic_up_down_button.controls.append(
                    TextButton(
                        data = "返回",
                        content = Container(
                            content = Column(
                                controls = [
                                    Text(value = f"返回", size = 20),
                                ],
                                alignment = MainAxisAlignment.CENTER,
                            ),
                        ),
                        on_click = lambda e:topic_up_down_button_clicked(e)
                    )
                )
                self.topic_up_down_button.controls.append(
                    TextButton(
                        data = "上一題",
                        content = Container(
                            content = Column(
                                controls = [
                                    Text(value = f"上一題", size = 20),
                                ],
                                alignment = MainAxisAlignment.CENTER,
                            ),
                        ),
                        disabled = True,
                        on_click = lambda e:topic_up_down_button_clicked(e)
                    )
                )
                self.topic_up_down_button.controls.append(
                    TextButton(
                        data = "下一題",
                        content = Container(
                            content = Column(
                                controls = [
                                    Text(value = f"下一題", size = 20),
                                ],
                                alignment = MainAxisAlignment.CENTER,
                            ),
                        ),
                        disabled = True,
                        on_click = lambda e:topic_up_down_button_clicked(e)
                    )
                )
                if self.index-1 > 0:
                    self.topic_up_down_button.controls[1].disabled = False

                if self.index < len(self.topic_data):
                    self.topic_up_down_button.controls[2].disabled = False

                # 建立修改內容
                def topic_edit_back_clicked(e):
                    self.topic_information_main.visible = True # 顯示第三區塊
                    self.topic_edit_main.visible = False # 隱藏第四區塊
                    page.update()
                def topic_edit_submit_clicked(e):
                    for index, _ in enumerate([self.topic_text_edit, self.topic_img_edit, self.topic_answer_edit, self.topic_option1_edit, self.topic_option2_edit, self.topic_option3_edit, self.topic_option4_edit]):
                        cloumn = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
                        self.wb["{}{}".format(cloumn[index], self.index)].value = f"{_.value}"
                    self.excel.save("exam_data.xlsx")
                    self.choice.visible = True # 顯示第一區塊
                    self.topic.visible = False # 隱藏第二區塊
                    self.topic_information_main.visible = False # 隱藏第三區塊
                    self.edit_content.visible = False # 隱藏修改內容選項
                    self.topic_edit_main.visible = False # 隱藏第四區塊
                    write_list()
                    page.update()
                topic_edit_back = TextButton(
                    content = Container(
                        content = Column(
                            controls = [
                                Text(value = f"返回", size = 20),
                            ],
                            alignment = MainAxisAlignment.CENTER,
                        ),
                    ),
                    on_click = lambda e:topic_edit_back_clicked(e)
                )
                topic_edit_submit = TextButton(
                    content = Container(
                        content = Column(
                            controls = [
                                Text(value = f"送出修改", size = 20),
                            ],
                            alignment = MainAxisAlignment.CENTER,
                        ),
                    ),
                    on_click = lambda e:topic_edit_submit_clicked(e)
                )
                for _ in [self.topic_text_edit, self.topic_img_edit, self.topic_answer_edit, self.topic_option1_edit, self.topic_option2_edit, self.topic_option3_edit, self.topic_option4_edit]:
                    self.topic_edit.controls.append(_)
                for _ in [topic_edit_back, topic_edit_submit]:
                    self.topic_edit_button.controls.append(_)
                page.update()
            def topic_list_clicked(e): # 點選題目
                try:
                    self.index = e.control.data # row的位置
                    self.topic.visible = False # 隱藏第二區塊
                    self.topic_information_main.visible = True # 顯示第三區塊
                    self.edit_content.visible = True # 顯示修改內容選項
                    content = self.topic_data[self.index-1] # 讀取題目內容
                    # print(content)
                    create_topic_information(content)
                except TypeError:
                    # print("TypeERROR")
                    self.choice.visible = True # 顯示第一區塊
                    self.topic.visible = False # 隱藏第二區塊
                    self.topic_information_main.visible = False # 隱藏第三區塊
                    self.edit_content.visible = False # 隱藏修改內容按鍵
                    self.topic_edit_main.visible = False # 隱藏第四區塊
                    write_list() # 刷新工作表選擇
                    warning_info("無法讀取舊版資料")
                page.update()
     
            # 創建題目和上下頁
            def page_button_clicked(e): # 按上、下頁
                # print(e.control.data)
                if e.control.data == "上一頁":
                    self.topic_index -= 100
                    create_topic()
                if e.control.data == "下一頁":
                    create_topic()
            def create_topic(end_index:int = 0):
                # 清空元件
                self.topic_list.controls = []
                self.page_button.controls = []
                # 判斷是否有上一頁、下一頁
                self.page_button.controls.append(
                    TextButton(
                        data = "上一頁",
                        content = Container(
                            content = Column(
                                controls = [
                                    Text(value = f"上一頁", size = 20),
                                ],
                                alignment = MainAxisAlignment.CENTER,
                            ),
                        ),
                        disabled = True,
                        on_click = lambda e:page_button_clicked(e)
                    )
                )
                self.page_button.controls.append(
                    TextButton(
                        data = "下一頁",
                        content = Container(
                            content = Column(
                                controls = [
                                    Text(value = f"下一頁", size = 20),
                                ],
                                alignment = MainAxisAlignment.CENTER,
                            ),
                        ),
                        disabled = True,
                        on_click = lambda e:page_button_clicked(e)
                    )
                )
                if self.topic_index > 0: # 上一頁
                    self.page_button.controls[0].disabled = False
                if self.topic_index + 50 < len(self.topic_data): # 下一頁
                    self.page_button.controls[1].disabled = False
                page.update()
                # 每一題的按鈕
                for index, _ in enumerate(self.topic_data[self.topic_index:self.topic_index+50]):
                    if _[0] != None:
                        self.topic_list.controls.append(
                            TextButton(
                                data = self.topic_index+index+1, # 第幾題(row)來取得excel資料
                                content = Container(
                                    content = Column(
                                        controls = [
                                            Text(value = "{}.({})：{}".format(self.topic_index+index+1, _[2], _[0].split("!n")[0][4:]), size = 18),
                                        ],
                                        alignment = MainAxisAlignment.CENTER,
                                    ),
                                ),
                                on_click = lambda e:topic_list_clicked(e)
                            )
                        )
                        # print(f"{self.topic_index+index+1}.{_[0]}")
                self.topic_index += 50
                page.update()

            if self.choose_sheetname_text.value[3:] in self.excel.sheetnames: # 是否有輸入題庫
                self.choice.visible = False # 隱藏第一區塊
                # 導入題目至清單
                self.wb = self.excel[self.choose_sheetname_text.value[3:]]
                self.topic_data = get_values(self.wb) # 讀取excel工作表裡的每一行
                self.topic_index = 0 # 起始題號
                self.topic.controls[0].value = f"目前題庫：{self.choose_sheetname_text.value[3:]}，請選擇題目。" # 顯示目前的題庫
                create_topic() # 創建題目
                self.topic.visible = True # 顯示第二區塊
            else:
                warning_info("請確認選擇的題庫是否正確")
            page.update()


        # 區塊1：選擇題庫
        def load_excel(): # 抓取excel資料
            if self.excel != None:
                self.excel.close()
            self.excel = openpyxl.load_workbook("exam_data.xlsx") # 讀取excel檔
            self.choose_sheetname.items = []
            page.update()
        def write_list(): # 寫入工作表清單
            self.excel = None
            load_excel()
            for workbook in self.excel.sheetnames[1:]:
                self.choose_sheetname.items.append(PopupMenuItem(text = f"{workbook}", on_click = lambda e:choose_sheetname_clicked(e))) 
        def choose_sheetname_clicked(e):
            self.choose_sheetname_text.value = f"題庫：{e.control.text}"
            page.update()
        # 確認送出
        self.review_submit = TextButton(
            content = Container(
                content = Column(
                    controls = [
                        Text(value = "確認", size = 20),
                    ],
                    alignment=MainAxisAlignment.CENTER,
                    spacing = 5,
                ),
                padding = padding.all(10),
            ),
            on_click = review_submit_clicked)
        self.choose_sheetname_text = Text("題庫：", size = 20, weight = "bold")
        self.choose_sheetname = PopupMenuButton(
            items=[],
            icon = icons.CLEAR_ALL_OUTLINED,
        )
        write_list()     
        # 主區塊
        self.choice = Column(
            expand = True,
            visible = True,
            controls = [
                Text("請選擇要複習的題庫", size = 20, weight = "bold"),
                Row(
                    controls = [
                        self.choose_sheetname_text,
                        self.choose_sheetname,
                    ],
                ),
                self.review_submit,
            ]
        )
        
        # 區塊2：題目按鈕清單
        # 題目按鈕清單
        self.topic_list = Column(
            controls = [],
        )
        # 上下頁按鈕區
        self.page_button = Row(
            controls = [],
        )   
        # 主區塊
        self.topic = Column(
            expand = True,
            visible = False,
            controls = [
                Text("目前題庫：，請選擇題目。", size = 20, weight = "bold"),
                Column(
                    expand = True,
                    scroll = "ADAPTIVE",
                    controls = [self.topic_list],
                ),
                Divider(height=1),
                self.page_button,
            ],
        )
        
        # 區塊3：顯示選中題目內容
        self.topic_up_down_button = Row(
            controls = [],
        )
        self.topic_information = Column(
            expand = True,
            scroll = "AUTO",
            controls = [],
        )
        self.topic_information_main = Column(
            expand = True,
            visible = False,
            controls = [
                self.topic_information,
                Divider(height=1),
                self.topic_up_down_button,
            ],
        )

        # 區塊4：修改內容
        self.topic_edit = Column(
            expand = True,
            scroll = "AUTO",
            controls = [],
        )
        self.topic_edit_button = Row(
            controls = []
        )
        self.topic_edit_main = Column(
            expand = True,
            visible = False,
            spacing = 10,
            controls = [
                self.topic_edit,
                Divider(height=1),
                self.topic_edit_button,
            ],
        )

    def view(self, page: Page, params: Params, basket: Basket):
        return View(
            controls = [
                Row(
                    expand = True,
                    controls = [
                        Row(
                            width = 100,
                            controls = [rail_item(page)],
                        ),
                        VerticalDivider(width=1),
                        Column(
                            expand = True,
                            alignment=MainAxisAlignment.START,
                            controls = [
                                Row(
                                    controls = [
                                        Text("複習題庫題目", size=30, weight="bold"),
                                        self.restart,
                                        self.edit_content,
                                    ]
                                ),
                                Divider(height=1),
                                self.choice,
                                self.topic,
                                self.topic_information_main,
                                self.topic_edit_main,
                            ],
                        ),
                    ],
                ),
            ],
        )